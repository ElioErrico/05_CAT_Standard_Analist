# helpers.py
import re, json
import os
from typing import Sequence, Any, List, Dict
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def _clean_cid_and_control_chars(text: str) -> str:
    text = re.sub(r'\(cid:\d+\)', '', text)
    return re.sub(r'[\x00-\x1f\x7f-\x9f\u200b-\u200f\u2028-\u202e]', '', text)

# --- Util per estrazione JSON robusta (riuso inline) ---
def _extract_json_object(text: str):
    fence = re.search(r"```(?:json|python)?\s*([\s\S]*?)```", text, flags=re.IGNORECASE)
    candidates = []
    if fence:
        candidates.append(fence.group(1).strip())
    candidates.append(text)

    for cand in candidates:
        try:
            obj = json.loads(cand)
            if isinstance(obj, list):
                return {"rows": obj}
            return obj
        except Exception:
            # prova bounding {} o [] nel fallback
            s, e = cand.find("{"), cand.rfind("}")
            if s != -1 and e != -1 and e > s:
                try:
                    obj = json.loads(cand[s:e+1])
                    if isinstance(obj, list):
                        return {"rows": obj}
                    return obj
                except Exception:
                    pass
            s, e = cand.find("["), cand.rfind("]")
            if s != -1 and e != -1 and e > s:
                try:
                    arr = json.loads(cand[s:e+1])
                    return {"rows": arr if isinstance(arr, list) else [arr]}
                except Exception:
                    pass
    raise ValueError("Impossibile estrarre un JSON valido dalla risposta del modello.")

# ----------------- Normalizzazione nomi colonna (spazi/NBSP/alias) -----------------
_WS = re.compile(r"\s+")

def _norm_key(s: str) -> str:
    if s is None:
        return ""
    s = str(s).replace("\xa0", " ")   # NBSP -> spazio normale
    s = _WS.sub(" ", s).strip()       # collassa spazi multipli
    return s

def _rename_to_canonical(
    df: pd.DataFrame,
    final_columns: Sequence[str],
    aliases: Dict[str, Sequence[str]] | None = None
) -> pd.DataFrame:
    """
    Rinomina le colonne correnti ai nomi canonici in final_columns:
    - matching case-insensitive su chiavi normalizzate (spazi multipli → singolo, NBSP rimossi)
    - supporta alias espliciti per varianti note
    """
    # base: tutte le colonne canoniche
    norm_to_canon = {_norm_key(c): c for c in final_columns}

    # alias comuni per "Chapter/Paragraph  No." (due spazi)
    default_aliases: Dict[str, List[str]] = {
        "Chapter/Paragraph  No.": [
            "Chapter/Paragraph No.",
            "Chapter / Paragraph No.",
            "Chapter / Paragraph  No.",
            "Chapter-Paragraph No.",
        ],
    }
    if aliases:
        # mergia alias extra
        for canon, alist in aliases.items():
            default_aliases.setdefault(canon, []).extend(alist)

    # registra alias normalizzati
    for canon, alist in default_aliases.items():
        norm_to_canon[_norm_key(canon)] = canon
        for a in alist:
            norm_to_canon[_norm_key(a)] = canon

    # costruisci rename map
    rename_map = {}
    for col in df.columns:
        key = _norm_key(col)
        if key in norm_to_canon and col != norm_to_canon[key]:
            rename_map[col] = norm_to_canon[key]

    if rename_map:
        df = df.rename(columns=rename_map)
    return df

# ----------------- DataFrame normalize + Excel writer -----------------
def normalize_rows_to_dataframe(
    rows: List[dict],
    required_keys: Sequence[str],
    extra_empty_columns: Sequence[str],
    final_columns: Sequence[str],
    dedup_keys: Sequence[str] = (
        "Chapter/Paragraph  No.",
        "Requirement/Standard Description",
        "Regulatory References",
    ),
) -> pd.DataFrame:
    """
    Converte 'rows' in DataFrame con:
    - json_normalize
    - rinomina colonne ai nomi canonici (gestione spazi/NBSP/alias)
    - garanzia colonne richieste/extra
    - dedup su dedup_keys
    - ordine colonne final_columns
    """
    df = pd.json_normalize(rows) if rows else pd.DataFrame()

    # 1) rinomina ai nomi canonici (es. "Chapter/Paragraph No." -> "Chapter/Paragraph  No.")
    df = _rename_to_canonical(df, final_columns)

    # 2) garantisce colonne richieste/extra/finali
    for col in required_keys:
        if col not in df.columns:
            df[col] = ""
    for col in extra_empty_columns:
        if col not in df.columns:
            df[col] = ""
    for col in final_columns:
        if col not in df.columns:
            df[col] = ""

    # 3) dedup su colonne chiave (se presenti)
    existing = [c for c in dedup_keys if c in df.columns]
    if existing:
        df = df.drop_duplicates(subset=existing, keep="first").reset_index(drop=True)

    # 4) ordine finale
    df = df.reindex(columns=final_columns)
    return df


def normalize_rows_and_write_excel(
    rows: List[dict],
    required_keys: Sequence[str],
    extra_empty_columns: Sequence[str],
    final_columns: Sequence[str],
    file_path: str,
    sheet_name: str = "Sheet1",
) -> pd.DataFrame:
    """
    Normalizza + scrive Excel con formattazione colonne/word-wrap:
      - Col 1: 20; Col 2: 20; Col 3: 60; resto: 20
      - Tutte le celle wrap_text e vertical='top'
    """
    df = normalize_rows_to_dataframe(rows, required_keys, extra_empty_columns, final_columns)

    os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # larghezze colonne
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = 60 if col_idx == 3 else 20

        # wrap text + vertical top
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=(cell.alignment.horizontal if cell.alignment else None),
                    vertical="top",
                    wrap_text=True
                )

    return df

# ----------------- utility testo -----------------
def split_text_into_n_parts(text: str, n: int) -> list[str]:
    if n <= 1 or not text:
        return [text]
    length = len(text)
    approx = max(1, length // n)
    parts, start = [], 0
    for _ in range(n - 1):
        cut = start + approx
        newline_pos = text.rfind("\n", start, min(length, cut + approx // 2))
        if newline_pos != -1 and newline_pos > start:
            cut = newline_pos + 1
        parts.append(text[start:cut])
        start = cut
    parts.append(text[start:])
    return parts
